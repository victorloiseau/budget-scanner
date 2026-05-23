import streamlit as st
import anthropic
import base64
import json
import re
import os
import io
from PIL import Image
import openpyxl
from datetime import date

# ── Config ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Tickets → Budget",
    page_icon="🧾",
    layout="centered",
)

SHEET_NAME = "Tracker mensuel"
REEL_COL   = 4
NOTE_COL   = 6

CATEGORIES = [
    ("Epicerie",                  5),
    ("Restaurants / cafeteria",   6),
    ("Cafe / boissons",           7),
    ("STM (abonnement etudiant)", 8),
    ("Telephone canadien",        9),
    ("Internet",                 10),
    ("Hygiene & soin",           11),
    ("Pharmacie",                12),
    ("Livres & materiel etudes", 13),
    ("Impression / fournitures", 14),
    ("Sorties & loisirs",        15),
    ("Streaming",                16),
    ("Sport / gym",              17),
    ("Frais etudiants HEC",      18),
    ("Imprevus",                 19),
]
CAT_NAMES = [n for n, _ in CATEGORIES]
CAT_ROW   = {n: r for n, r in CATEGORIES}

# ── API key ───────────────────────────────────────────────────────────────────
def get_api_key() -> str | None:
    # 1. Streamlit secrets (déploiement cloud)
    if "ANTHROPIC_API_KEY" in st.secrets:
        return st.secrets["ANTHROPIC_API_KEY"]
    # 2. Variable d'environnement (local)
    if "ANTHROPIC_API_KEY" in os.environ:
        return os.environ["ANTHROPIC_API_KEY"]
    return None


# ── Analyse du ticket via Claude Vision ───────────────────────────────────────
def analyze_receipt(image: Image.Image, api_key: str) -> dict:
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=90)
    b64 = base64.standard_b64encode(buf.getvalue()).decode()

    client = anthropic.Anthropic(api_key=api_key)

    prompt = f"""Analyse ce ticket de caisse et réponds UNIQUEMENT avec un objet JSON valide, sans markdown, sans explication.

Format attendu :
{{
  "store": "nom exact du magasin ou commerce",
  "total": 12.34,
  "date": "{date.today().isoformat()}",
  "category": "categorie exacte parmi la liste ci-dessous",
  "items_summary": "résumé bref des achats en 1 ligne"
}}

Catégories disponibles (choisis la plus adaptée, texte exact) :
- Epicerie
- Restaurants / cafeteria
- Cafe / boissons
- Pharmacie
- Hygiene & soin
- Livres & materiel etudes
- Sorties & loisirs
- Sport / gym
- Streaming
- Impression / fournitures
- Imprevus

Si le total n'est pas lisible sur le ticket, mets null.
Si la date n'est pas lisible, mets "{date.today().isoformat()}".
Réponds uniquement avec le JSON, sans rien d'autre."""

    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=400,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
                },
                {"type": "text", "text": prompt},
            ],
        }],
    )

    raw = msg.content[0].text.strip()
    # Nettoie les éventuels blocs markdown
    raw = re.sub(r"^```(?:json)?\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)
    return json.loads(raw)


# ── Excel ─────────────────────────────────────────────────────────────────────
def read_current(path: str, row: int) -> float:
    wb  = openpyxl.load_workbook(path)
    val = wb[SHEET_NAME].cell(row=row, column=REEL_COL).value
    wb.close()
    return float(val) if val else 0.0


def save_expense(path: str, row: int, amount: float, note: str):
    wb  = openpyxl.load_workbook(path)
    ws  = wb[SHEET_NAME]
    cur = ws.cell(row=row, column=REEL_COL).value or 0
    ws.cell(row=row, column=REEL_COL).value = round(float(cur) + amount, 2)
    if note:
        existing = ws.cell(row=row, column=NOTE_COL).value or ""
        ws.cell(row=row, column=NOTE_COL).value = (existing + " | " if existing else "") + note
    wb.save(path)
    wb.close()


# ── UI ────────────────────────────────────────────────────────────────────────
st.title("🧾 Tickets de caisse → Budget Montréal")
st.caption("Prends une photo de ton ticket, la dépense s'ajoute automatiquement dans ton budget.")

# Sidebar : config
with st.sidebar:
    st.header("Configuration")

    api_key = get_api_key()
    if not api_key:
        api_key = st.text_input("Clé API Anthropic", type="password",
                                help="console.anthropic.com → API Keys")
    else:
        st.success("Clé API chargée")

    st.divider()
    st.subheader("Fichier Excel")
    excel_mode = st.radio("Fichier budget", ["Uploader le fichier", "Chemin local (PC uniquement)"])

    excel_path  = None
    excel_bytes = None

    if excel_mode == "Uploader le fichier":
        uploaded_excel = st.file_uploader("budget_montreal_v2.xlsx", type=["xlsx"])
        if uploaded_excel:
            excel_bytes = uploaded_excel.read()
            st.success("Fichier chargé")
    else:
        default_path = r"C:\Users\victo\Desktop\budget_montreal_v2.xlsx"
        excel_path   = st.text_input("Chemin", value=default_path)
        if os.path.exists(excel_path):
            st.success("Fichier trouvé")
        else:
            st.error("Fichier introuvable")

# Vérification prérequis
if not api_key:
    st.warning("Entre ta clé API Anthropic dans la barre latérale pour commencer.")
    st.info("Crée un compte gratuit sur **console.anthropic.com** → API Keys → Create Key. Tu reçois 5 $ de crédit offert (~5 000 tickets gratuits).")
    st.stop()

if excel_mode == "Uploader le fichier" and not excel_bytes:
    st.info("Upload ton fichier Excel dans la barre latérale.")
    st.stop()
elif excel_mode == "Chemin local (PC uniquement)" and not os.path.exists(excel_path):
    st.error("Fichier Excel introuvable.")
    st.stop()

# Upload ticket
st.divider()
uploaded = st.file_uploader("📷 Dépose la photo de ton ticket ici", type=["jpg","jpeg","png","webp","heic"])

if not uploaded:
    st.markdown("""
**Astuce téléphone** : prends la photo dans un endroit bien éclairé, ticket bien à plat.
Le ticket peut être en français ou en anglais.
    """)
    st.stop()

img = Image.open(uploaded)
st.image(img, caption="Ticket reçu", use_column_width=True)

if st.button("🔍 Analyser le ticket", type="primary", use_container_width=True):
    with st.spinner("Claude lit ton ticket..."):
        try:
            result = analyze_receipt(img, api_key)
            st.session_state["result"]   = result
            st.session_state["analyzed"] = True
        except json.JSONDecodeError:
            st.error("Le ticket n'a pas pu être analysé. Essaie avec une photo plus nette.")
            st.session_state["analyzed"] = False
        except anthropic.AuthenticationError:
            st.error("Clé API invalide. Vérifie ta clé dans la barre latérale.")
            st.session_state["analyzed"] = False
        except Exception as e:
            st.error(f"Erreur : {e}")
            st.session_state["analyzed"] = False

if not st.session_state.get("analyzed"):
    st.stop()

result = st.session_state["result"]

# Résultats
st.divider()
st.subheader("Ce que Claude a lu sur ton ticket")

col_info, col_form = st.columns([1, 1])

with col_info:
    st.markdown(f"**Magasin :** {result.get('store', 'Non détecté')}")
    st.markdown(f"**Date :** {result.get('date', date.today().isoformat())}")
    if result.get("items_summary"):
        st.markdown(f"**Achats :** {result['items_summary']}")

with col_form:
    detected_total = result.get("total")
    if detected_total:
        amount = st.number_input("Montant ($ CA)", value=float(detected_total),
                                  min_value=0.01, step=0.01, format="%.2f")
    else:
        st.warning("Montant non détecté, entre-le manuellement.")
        amount = st.number_input("Montant ($ CA)", min_value=0.01, step=0.01, format="%.2f")

    detected_cat = result.get("category", "Imprevus")
    default_idx  = CAT_NAMES.index(detected_cat) if detected_cat in CAT_NAMES else len(CAT_NAMES)-1
    category     = st.selectbox("Catégorie", CAT_NAMES, index=default_idx)

note = st.text_input("Note (optionnel)", placeholder=f"ex: {result.get('store','Ticket')} du {result.get('date', date.today().isoformat())}")

# Aperçu de l'impact
row = CAT_ROW[category]
if excel_bytes:
    tmp_path = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Temp", "budget_tmp.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(excel_bytes)
    before = read_current(tmp_path, row)
else:
    before = read_current(excel_path, row)

after = round(before + amount, 2)
st.markdown(f"""
| Poste | Avant | Après ajout |
|---|---|---|
| **{category}** | {before:.2f} $ CA | **{after:.2f} $ CA** |
""")

# Enregistrement
if st.button("✅ Ajouter au budget", type="primary", use_container_width=True):
    try:
        if excel_bytes:
            save_expense(tmp_path, row, amount, note)
            with open(tmp_path, "rb") as f:
                updated = f.read()
            st.success(f"**{amount:.2f} $ CA** ajouté à **{category}** — nouveau total : **{after:.2f} $ CA**")
            st.download_button(
                "⬇️ Télécharger le budget mis à jour",
                data=updated,
                file_name="budget_montreal_v2.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            save_expense(excel_path, row, amount, note)
            st.success(f"**{amount:.2f} $ CA** ajouté à **{category}** — nouveau total : **{after:.2f} $ CA**")

        st.balloons()
        st.session_state["analyzed"] = False
    except PermissionError:
        st.error("Ferme le fichier Excel dans Excel, puis réessaie.")
    except Exception as e:
        st.error(f"Erreur : {e}")
